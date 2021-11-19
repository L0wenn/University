"""
TODO: Enter the catalogue, step through all categories
until we find parts table and blueprint(bp), save the bp,
step through all parts in the table, save images of parts
and their info and write all of it into xlsx file
""" 

import json
import os
import shutil
from typing import Union
from urllib.parse import urljoin
from zipfile import BadZipFile

import openpyxl
from bs4 import BeautifulSoup as bs
from selenium import webdriver
from selenium.common.exceptions import (NoSuchElementException,
                                        TimeoutException, WebDriverException)
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.firefox.webdriver import WebDriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

URL = "https://dealler.ru/katalog"

with open("ignored.json", "r") as f:
    data = f.read()
    ignore_links = json.loads(data)["ignored"]


def parse(url:str, driver: WebDriver):
    """
    TODO: Function must step through the catalogue
    recursively untill it finds div with class products-view
    """
    if url in ignore_links:
        return

    page = connect_to(url, driver)
    soup = bs(page.page_source, "html.parser")

    if soup.find("div", class_="products-view") is None:
        div = soup.find_all("div", class_="category")

        for category in div:
            a = category.find("a")

            absolute_link = urljoin(URL, a["href"])
            if absolute_link in ignore_links:
                continue
            
            parse(absolute_link, page)
            add_link_to_ignored(absolute_link)
    else:
        blueprint = driver.current_url
        save_part(blueprint, page)
        add_link_to_ignored(blueprint)

def save_part(url: str, driver: WebDriver):
    # Download blueprints first
    if url in ignore_links:
        return

    soup = bs(driver.page_source, "html.parser")
    bp_div = soup.find("div", class_="category_description")
    bp_img = bp_div.find("img")
    img_url = URL.replace("/katalog", bp_img["src"])
    download_image(img_url, driver)

    # Next download all parts from tables
    products = soup.find_all("div", class_="product-name")
    for product in products:
        a = product.find("a")

        if a == None:
            continue

        absolute_link = urljoin(URL, a["href"])
        if absolute_link in ignore_links:
            continue

        page = connect_to(absolute_link, driver)
        if not page:
            continue 

        soup = bs(page.page_source, "html.parser")
        img_div = soup.find("div", class_="main-image")

        try:
            img_url = img_div.find("a")["href"]
        except AttributeError:
            continue

        download_image(img_url, page)
        write_to_xls(absolute_link, img_url, page)
        add_link_to_ignored(absolute_link)

def write_to_xls(url: str, image, driver: WebDriver):
    try:
        wb = openpyxl.load_workbook("data.xlsx")
    except BadZipFile:
        os.remove("data.xlsx")
        os.rename("recovery_data.xlsx", "data.xlsx")
        wb = openpyxl.load_workbook("data.xlsx")

    page = wb.active
    filename = image.split("/")[-1]
    filename = filename if not filename.startswith("image not available") else "chinatown.jpg"
    
    soup = bs(driver.page_source, "html.parser")

    product = soup.find("div", class_="product-area")
    manufacturer = product.find("div", class_="manufacturer")
    name = soup.find("h1")
    desc = product.find("div", "product-short-description")
    number = soup.find("b").text

    ul = soup.find("ul", class_="breadcrumb")
    a = ul.find_all("a")
    path = []

    for el in a:
        path.append(el.find("span").text)

    path = "/".join(path)

    shutil.copy("data.xlsx", "recovery_data.xlsx")
    page.append([filename, manufacturer.text, number, name.text, desc.text.replace("\n", "").replace("\t", ""), path.replace("Каталоги/", "")])
    wb.save("data.xlsx")
    wb.close()
    
def download_image(url: str, driver: WebDriver):
    filename = url.split("/")[-1].replace(".jpg", ".png")
    path = f"images/{filename}"

    if filename == "image not available.png":
        return
    
    with open(path, "wb") as f:
        try:
            img = driver.find_element_by_xpath('/html/body/section/section[3]/div/div/div/div[2]/div/div[3]/div/div/div[2]/div[2]/div[1]/div/a/img')
        except NoSuchElementException:
            img = driver.find_element_by_xpath('/html/body/section/section[3]/div/div/div/div[2]/div/div[3]/div/div/div[2]/div[1]/p/img')
        
        try:
            f.write(img.screenshot_as_png)
        except WebDriverException:
            pass

def connect_to(url: str, driver: WebDriver) -> Union[WebDriver, bool]:
    try:
        wait = WebDriverWait(driver, 10)
        driver.get(url)
        if check_400(driver):
            return False
            
        wait.until(EC.presence_of_all_elements_located((By.TAG_NAME, "a")))

        return driver
    except TimeoutException:
        print("Failed to connect to:", url, "\nRetrying...")
        connect_to(url, driver)

def add_link_to_ignored(url: str):
    if url in ignore_links:
        return

    ignore_links.append(url)
    d = {"ignored": ignore_links}

    with open("ignored.json", "w") as f:
        json.dump(d, f, indent=2, separators=(",", ": "))

def check_400(driver: WebDriver) -> bool:
    soup = bs(driver.page_source, "html.parser")
    center = soup.find("center")
    try:
        h1 = center.find("h1").text
    except AttributeError:
        return False

    return True if h1.lower() == "400 bad request" else False


if __name__ == "__main__":
    if not os.path.exists("data.xlsx"):
        headers = [
        "Название изображения",
        "Название производителя",
        "Номер детали",
        "Название детали",
        "Развёрнутое описание детали",
        "category"
        ]
        
        wb = openpyxl.Workbook()
        page = wb.active
        page.append(headers)
        wb.save("data.xlsx")
        wb.close()

    opts = Options()
    ff_driver = os.getcwd() + "/geckodriver"
    driver = webdriver.Firefox(options=opts, executable_path=ff_driver)
    parse(URL, driver)
    driver.quit()
