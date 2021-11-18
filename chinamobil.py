"""
TODO: Enter the catalogue, step through all categories
until we find parts table and blueprint(bp), save the bp,
step through all parts in the table, save images of parts
and their info and write all of it into xlsx file
""" 

import json
import os
import shutil
from typing import Union
from urllib.parse import urljoin
from zipfile import BadZipFile

import openpyxl
from bs4 import BeautifulSoup as bs
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.firefox.webdriver import WebDriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

URL = "https://чинамобил.рф/magazin/folder/zapchasti-katalog-geely"

with open("ignored.json", "r") as f:
    data = f.read()
    ignore_links = json.loads(data)["ignored"]


def parse(url:str, driver: WebDriver):
    if url in ignore_links:
        return

    page = connect_to(url, driver)
    soup = bs(page.page_source, "html.parser")
    
    # I'm too fucking sleepy to deal with this one
    wrap_img_main = soup.find("div", class_="wrap-img-main")
    try:
        a = wrap_img_main.find("a")
    except Exception:
        a = None
    
    if soup.find("div", class_="win") or soup.find("form", action="/magazin?mode=cart&action=add") or a:
        absolute_link = None

        try:
            save_part(page.current_url, page)
        except AttributeError:
            return

        try:
            for folder in soup.find_all("div", class_="block-main-img"):
                a = folder.find("a")
                img = a.find("img")["src"]
                
                if "spacer.gif" in img:
                    continue

                absolute_link = urljoin(URL, a["href"])
                if absolute_link in ignore_links:
                    continue

                parse(absolute_link, page)
        except Exception:
            pass

        add_link_to_ignored(absolute_link)
    elif soup.find("div", class_="block-main-img") is None:
        if soup.find("div", class_="wrap-img-main"):
            return

        categories = soup.find("nav", class_="kategor2")
        lis = categories.find_all("li")

        for li in lis:
            a = li.find("a")["href"]

            absolute_link = urljoin(URL, a)
            if absolute_link in ignore_links:
                continue

            parse(absolute_link, page)
            add_link_to_ignored(absolute_link)
    else:
        for found_div in soup.find_all("div", class_="block-main-img"):
            a = found_div.find("a")["href"]

            absolute_link = urljoin(URL, a)
            if absolute_link in ignore_links:
                continue

            parse(absolute_link, page)
            add_link_to_ignored(absolute_link)

def save_part(url: str, driver: WebDriver):
    if url in ignore_links:
        return

    write_to_xls(url, driver)

def write_to_xls(url: str, driver: WebDriver):
    try:
        wb = openpyxl.load_workbook("data.xlsx")
    except BadZipFile:
        os.remove("data.xlsx")
        os.rename("recovery_data.xlsx", "data.xlsx")
        wb = openpyxl.load_workbook("data.xlsx")

    page = wb.active
    soup = bs(driver.page_source, "html.parser")

    part_img_name = ""
    part_NO = ""
    part_name = ""
    category = ""

    path = soup.find("div", class_="site-path").text.replace("Главная \ ", "")
    manufacturer = path.split(" \ ")[0]
    table = soup.find("form", action="/magazin?mode=cart&action=add")

    if table is None:
        desc = soup.find("div", class_="f_desc")
        p = desc.find("p")
        parts = p.get_text(strip=True, separator="\n").splitlines()

        for part in parts:
            part = part.replace("\xa0", " ")

            part_NO = part.split(" ")[-1]
            part_img_name = part_NO + ".jpg"
            part_name = " ".join(part.split(" ")[:-1])
            category = path.replace("Главная \ ", "").replace(" \ ", "/")
            
            page.append([part_img_name, manufacturer, part_NO, part_name, part_name, category])
            wb.save("data.xlsx")
    else:
        tbody = table.find("tbody")
        oddtr = tbody.find_all("tr", class_="odd")
        eventr = tbody.find_all("tr", class_="even")
        tr = oddtr + eventr
        
        for element in tr:
            text = element.get_text("|", strip=True)
            text = text.split("|")
            
            part_NO = text[1]
            part_img_name = part_NO + ".jpg"
            part_name = text[0]
            category = path.replace("Главная \ ", "").replace(" \ ", "/")
            
            page.append([part_img_name, manufacturer, part_NO, part_name, part_name, category])
            wb.save("data.xlsx")
    
    shutil.copy("data.xlsx", "recovery_data.xlsx")     
    wb.close()

def connect_to(url: str, driver: WebDriver) -> Union[WebDriver, bool]:
    try:
        wait = WebDriverWait(driver, 10)
        driver.get(url)
        wait.until(EC.presence_of_all_elements_located((By.TAG_NAME, "a")))

        return driver
    except TimeoutException:
        print("Failed to connect to:", url, "\nRetrying...")
        connect_to(url, driver)

def add_link_to_ignored(url: str):
    if url in ignore_links:
        return

    ignore_links.append(url)
    d = {"ignored": ignore_links}

    with open("ignored.json", "w") as f:
        json.dump(d, f, indent=2, separators=(",", ": "))

def check_400(driver: WebDriver) -> bool:
    soup = bs(driver.page_source, "html.parser")
    center = soup.find("center")
    try:
        h1 = center.find("h1").text
    except AttributeError:
        return False

    return True if h1.lower() == "400 bad request" else False


if __name__ == "__main__":
    if not os.path.exists("data.xlsx"):
        headers = [
        "Название изображения",
        "Название производителя",
        "Номер детали",
        "Название детали",
        "Развернутое описание детали",
        "category"
        ]
        
        wb = openpyxl.Workbook()
        page = wb.active
        page.append(headers)
        wb.save("data.xlsx")
        wb.close()

    opts = Options()
    ff_driver = os.getcwd() + "/geckodriver"
    driver = webdriver.Firefox(options=opts, executable_path=ff_driver)
    parse(URL, driver)
    driver.quit()
